from openpyxl import Workbook, load_workbook
from excel import generate_report


def create_test_excel(file_name):

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "CATEGORY"
    ws["B1"] = "TICKET"

    ws.append(["IRREGULAR_TEST_FAILURE", "SIT-123"])
    ws.append(["REGULAR_BUILD_FAILURE", "SIT-456"])
    ws.append(["IRREGULAR_TEST_FAILURE", None])

    wb.save(file_name)


def test_output_file_created(tmp_path):

    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    create_test_excel(input_file)

    generate_report(input_file, output_file)

    assert output_file.exists()


def test_only_irregular_failures_present(tmp_path):

    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    create_test_excel(input_file)

    generate_report(input_file, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    categories = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        categories.append(row[0])

    assert categories == ["IRREGULAR_TEST_FAILURE", "IRREGULAR_TEST_FAILURE"]


def test_blank_ticket_replaced_with_null(tmp_path):

    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    create_test_excel(input_file)

    generate_report(input_file, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    tickets = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        tickets.append(row[1])

    assert "NULL" in tickets


def test_regular_build_failure_removed(tmp_path):

    input_file = tmp_path / "input.xlsx"
    output_file = tmp_path / "output.xlsx"

    create_test_excel(input_file)

    generate_report(input_file, output_file)

    wb = load_workbook(output_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        assert row[0] != "REGULAR_BUILD_FAILURE"

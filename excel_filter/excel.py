from copy import copy
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def generate_report(input_file, output_file):
    source_wb = load_workbook(input_file)
    source_ws = source_wb[source_wb.sheetnames[0]]

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Report"

    category_col = None
    ticket_col = None
    header_row = None

    for row in source_ws.iter_rows():
        for cell in row:
            if cell.value == "CATEGORY":
                category_col = cell.column
                header_row = cell.row

            if cell.value == "TICKET":
                ticket_col = cell.column

        if category_col and ticket_col:
            break

    for col_num in range(1, source_ws.max_column + 1):
        col_letter = get_column_letter(col_num)

        if source_ws.column_dimensions[col_letter].width:
            source_width = source_ws.column_dimensions[
                col_letter
            ].width

            new_ws.column_dimensions[
                col_letter
            ].width = source_width

    target_row = 1

    for col in range(1, source_ws.max_column + 1):
        source_cell = source_ws.cell(header_row, col)

        target_cell = new_ws.cell(
            row=target_row,
            column=col,
            value=source_cell.value,
        )

        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.number_format = source_cell.number_format

    target_row += 1

    for row_num in range(header_row + 1, source_ws.max_row + 1):
        category_value = source_ws.cell(
            row=row_num,
            column=category_col,
        ).value

        if category_value is None:
            continue

        if str(category_value).strip() != "IRREGULAR_TEST_FAILURE":
            continue

        for col in range(1, source_ws.max_column + 1):
            cell_value = source_ws.cell(
                row=row_num,
                column=col,
            ).value

            if (
                col == ticket_col
                and (
                    cell_value is None
                    or str(cell_value).strip() == ""
                )
            ):
                cell_value = "NULL"

            source_cell = source_ws.cell(row_num, col)

            target_cell = new_ws.cell(
                row=target_row,
                column=col,
                value=cell_value,
            )

            target_cell.alignment = copy(
                source_cell.alignment
            )
            target_cell.number_format = (
                source_cell.number_format
            )

        target_row += 1

    new_wb.save(output_file)

    return output_file


if __name__ == "__main__":
    input_file = (
        r"C:\Users\40054008\Downloads"
        r"\failuresClassifications_MGU22_03-08-2026-10-08-2026.xlsx"
    )

    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    output_file = (
        r"C:\Users\40054008\OneDrive - LTTS"
        rf"\git_work\GitHubActionsDemo\excel_filter"
        rf"\IRREGULAR_TEST_FAILURE_Report_{timestamp}.xlsx"
    )

    generate_report(
        input_file,
        output_file,
    )

    print("File created successfully!")
    print("Saved at:", output_file)